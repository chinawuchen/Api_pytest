import openpyxl
import sys
import os
from openpyxl import workbook

base_path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
sys.path.append(base_path)

from Config.config import HandleInit


"""该模块主要是用来读取excel表格数据的"""


class HandExcel(object):

    def __init__(self):
        self.handle_ini = HandleInit()

    # 加载excel
    def load_excel(self):
        open_excel = openpyxl.load_workbook(
            base_path + self.handle_ini.get_value("case_path", "path"))
        return open_excel

    # 加载某个sheet的所有内容
    def get_sheet_data(self, index=None):
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    # 获取某一个单元格内容
    def get_cell_value(self, row, cols):
        data = self.get_sheet_data().cell(row=row, column=cols).value
        return data

    # 获取行数
    def get_rows(self):
        row = self.get_sheet_data().max_row
        return row

    # 获取某一行的内容
    def get_rows_value(self, row):
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        if row_list[0] != None:
            return row_list

    # 获取某一列的数据
    def get_columns_value(self, key=None):
        columns_list = []
        if key == None:
            key = 'A'
        for i in self.get_sheet_data()[key]:
            columns_list.append(i.value)
        return columns_list

    # 获取行号
    def get_rows_number(self, case_id):
        num = 1
        for i in self.get_columns_value():
            if case_id == i:
                return num
            num = num + 1
        return num

    # 写入数据
    def excel_write_data(self, row, cols, value):
        wb = self.load_excel()
        wr = wb.active # 默认第1个sheet
        wr.cell(row, cols, value)
        wb.save(base_path + self.handle_ini.get_value("case_path", "path"))

    # 获取excel里面所有的数
    def get_excel_data(self):
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_rows_value(i + 2))
        return data_list


if __name__ == "__main__":
    handle = HandExcel()
    # print(handle.load_excel())
    # print(handle.get_sheet_data())
    # print(handle.get_cell_value(8, 8))
    # print(handle.get_rows())
    # print(handle.get_rows_value(1))
    # print(handle.get_columns_value())
    # print(handle.get_rows_number("Hub_0001"))
    # print(handle.get_excel_data())
